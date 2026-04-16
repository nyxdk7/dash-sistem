import openpyxl


def limpar_texto(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def limpar_numero(valor):
    if valor is None or valor == '':
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    texto = texto.replace('R$', '').replace(' ', '')

    # padrão brasileiro: 1.234,56
    if ',' in texto:
        texto = texto.replace('.', '').replace(',', '.')

    try:
        return float(texto)
    except ValueError:
        return 0.0


def extrair_cabecalho_padrao(ws):
    cabecalho = {}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True):
        linha = [limpar_texto(c) for c in row]

        for i, valor in enumerate(linha):
            valor_lower = valor.lower()

            if valor_lower == 'processo:' and i + 1 < len(linha):
                cabecalho['processo'] = linha[i + 1]
            elif valor_lower == 'contrato:' and i + 1 < len(linha):
                cabecalho['contrato'] = linha[i + 1]
            elif valor_lower == 'contratada:' and i + 1 < len(linha):
                cabecalho['contratada'] = linha[i + 1]
            elif valor_lower == 'rodovia:' and i + 1 < len(linha):
                cabecalho['rodovia'] = linha[i + 1]
            elif valor_lower == 'trecho:' and i + 1 < len(linha):
                cabecalho['trecho'] = linha[i + 1]
            elif valor_lower == 'sub-trecho:' and i + 1 < len(linha):
                cabecalho['sub_trecho'] = linha[i + 1]
            elif valor_lower == 'segmento:' and i + 1 < len(linha):
                cabecalho['segmento'] = linha[i + 1]
            elif valor_lower == 'medição:' and i + 1 < len(linha):
                cabecalho['medicao'] = linha[i + 1]
            elif valor_lower == 'período:' and i + 1 < len(linha):
                cabecalho['periodo'] = linha[i + 1]
            elif valor_lower == 'período acum.:' and i + 1 < len(linha):
                cabecalho['periodo_acumulado'] = linha[i + 1]

    return cabecalho


def detectar_tipo_aba(ws):
    textos = []

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        for cell in row:
            if cell is not None:
                textos.append(str(cell).strip().lower())

    conteudo = ' '.join(textos)

    # modelo principal da sua planilha
    if 'código' in conteudo and 'preço unitário' in conteudo and 'financeiro' in conteudo:
        return 'contrato'

    # iluminação
    if 'descrição do produto' in conteudo and 'preço total' in conteudo:
        return 'iluminacao'

    # erosão / memória de cálculo
    if 'especificação dos serviços' in conteudo and 'volume' in conteudo:
        return 'erosao'

    # CBUQ por memória geométrica
    if 'relação do cbuq executado' in conteudo or (
        'comprimento' in conteudo and 'largura' in conteudo and 'espessura' in conteudo
    ):
        return 'cbuq'

    return 'desconhecido'


def localizar_linha_cabecalho_contrato(ws):
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        linha = [limpar_texto(c).lower() for c in row]

        if 'código' in linha and 'item' in linha:
            texto_linha = ' '.join(linha)
            if 'preço' in texto_linha or 'financeiro' in texto_linha or 'contrato' in texto_linha:
                return idx

    return None


def extrair_aba_contrato(ws):
    cabecalho = extrair_cabecalho_padrao(ws)
    itens = []

    linha_cabecalho = localizar_linha_cabecalho_contrato(ws)
    if not linha_cabecalho:
        return cabecalho, itens

    for row in ws.iter_rows(min_row=linha_cabecalho + 1, max_row=ws.max_row, values_only=True):
        # Estrutura observada na planilha:
        # 0 = código principal
        # 1 = código auxiliar/em alguns casos vazio
        # 2 = item
        # 3 = unid
        # 4 = contrato físico
        # 5 = contrato financeiro
        # 9 = preço unitário
        codigo_base = limpar_texto(row[0]) if len(row) > 0 else ''
        codigo_aux = limpar_texto(row[1]) if len(row) > 1 else ''
        descricao = limpar_texto(row[2]) if len(row) > 2 else ''
        unidade = limpar_texto(row[3]) if len(row) > 3 else ''
        quantidade = limpar_numero(row[4]) if len(row) > 4 else 0.0
        financeiro = limpar_numero(row[5]) if len(row) > 5 else 0.0
        preco_unitario = limpar_numero(row[9]) if len(row) > 9 else 0.0

        codigo = codigo_aux if codigo_aux else codigo_base

        # linha totalmente vazia
        if not codigo_base and not codigo_aux and not descricao and not unidade and quantidade == 0 and financeiro == 0:
            continue

        # ignora rodapé / totalizações finais
        descricao_lower = descricao.lower()
        if descricao_lower.startswith('total') or descricao_lower.startswith('subtotal'):
            continue

        # mantém linhas de grupo, mas sem quebrar
        item = {
            'aba': ws.title,
            'tipo_aba': 'contrato',
            'codigo': codigo,
            'descricao': descricao,
            'unidade': unidade,
            'quantidade': quantidade,
            'preco_unitario': preco_unitario,
            'preco_total': financeiro,
            'observacao': '',
            'marca': ''
        }

        # só adiciona se tiver alguma informação útil
        if item['codigo'] or item['descricao']:
            itens.append(item)

    return cabecalho, itens


def extrair_aba_iluminacao(ws):
    cabecalho = extrair_cabecalho_padrao(ws)
    itens = []

    for row in ws.iter_rows(min_row=15, max_row=ws.max_row, values_only=True):
        codigo = row[0] if len(row) > 0 else None
        descricao = limpar_texto(row[1]) if len(row) > 1 else ''
        marca = limpar_texto(row[2]) if len(row) > 2 else ''
        unidade = limpar_texto(row[3]) if len(row) > 3 else ''
        quantidade = limpar_numero(row[4]) if len(row) > 4 else 0.0
        preco_unitario = limpar_numero(row[5]) if len(row) > 5 else 0.0
        preco_total = limpar_numero(row[6]) if len(row) > 6 else 0.0

        if not codigo and not descricao:
            continue

        item = {
            'aba': ws.title,
            'tipo_aba': 'iluminacao',
            'codigo': str(codigo).strip() if codigo is not None else '',
            'descricao': descricao,
            'unidade': unidade,
            'quantidade': quantidade,
            'observacao': '',
            'marca': marca,
            'preco_unitario': preco_unitario,
            'preco_total': preco_total
        }

        itens.append(item)

    return cabecalho, itens


def extrair_aba_erosao(ws):
    cabecalho = extrair_cabecalho_padrao(ws)
    itens = []

    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, values_only=True):
        codigo = limpar_texto(row[1]) if len(row) > 1 else ''
        especificacao = limpar_texto(row[2]) if len(row) > 2 else ''
        unidade = limpar_texto(row[7]) if len(row) > 7 else ''
        volume = limpar_numero(row[8]) if len(row) > 8 else 0.0
        descricao_extra = limpar_texto(row[9]) if len(row) > 9 else ''

        if not codigo and not especificacao and not unidade and volume == 0:
            continue

        item = {
            'aba': ws.title,
            'tipo_aba': 'erosao',
            'codigo': codigo,
            'descricao': especificacao,
            'unidade': unidade,
            'quantidade': volume,
            'observacao': descricao_extra,
            'marca': '',
            'preco_unitario': 0.0,
            'preco_total': 0.0
        }

        if item['descricao'] or item['codigo']:
            itens.append(item)

    return cabecalho, itens


def extrair_aba_cbuq(ws):
    cabecalho = extrair_cabecalho_padrao(ws)
    itens = []

    for row in ws.iter_rows(min_row=18, max_row=ws.max_row, values_only=True):
        data = limpar_texto(row[1]) if len(row) > 1 else ''
        lado = limpar_texto(row[2]) if len(row) > 2 else ''
        km = limpar_texto(row[6]) if len(row) > 6 else ''
        comprimento = limpar_numero(row[7]) if len(row) > 7 else 0.0
        largura = limpar_numero(row[8]) if len(row) > 8 else 0.0
        espessura = limpar_numero(row[9]) if len(row) > 9 else 0.0
        area = limpar_numero(row[10]) if len(row) > 10 else 0.0
        volume = limpar_numero(row[11]) if len(row) > 11 else 0.0

        if not data and not lado and not km and area == 0 and volume == 0:
            continue

        item = {
            'aba': ws.title,
            'tipo_aba': 'cbuq',
            'codigo': '',
            'descricao': 'CBUQ executado',
            'unidade': 'm³',
            'quantidade': volume,
            'observacao': f'Data: {data} | Lado: {lado} | KM: {km} | Comp: {comprimento} | Larg: {largura} | Esp: {espessura} | Área: {area}',
            'marca': '',
            'preco_unitario': 0.0,
            'preco_total': 0.0
        }

        itens.append(item)

    return cabecalho, itens


def extrair_medicao(arquivo):
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    cabecalho_geral = None
    todos_itens = []

    for ws in wb.worksheets:
        tipo = detectar_tipo_aba(ws)

        if tipo == 'contrato':
            cabecalho, itens = extrair_aba_contrato(ws)
        elif tipo == 'iluminacao':
            cabecalho, itens = extrair_aba_iluminacao(ws)
        elif tipo == 'erosao':
            cabecalho, itens = extrair_aba_erosao(ws)
        elif tipo == 'cbuq':
            cabecalho, itens = extrair_aba_cbuq(ws)
        else:
            continue

        if not cabecalho_geral and cabecalho:
            cabecalho_geral = cabecalho

        todos_itens.extend(itens)

    return cabecalho_geral, todos_itens