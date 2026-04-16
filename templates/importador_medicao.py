import openpyxl


def texto(v):
    if v is None:
        return None
    return str(v).strip()


def numero(v, padrao=0.0):
    if v is None or v == "":
        return padrao
    try:
        return float(v)
    except (TypeError, ValueError):
        return padrao


def nome_aba_normalizado(nome):
    return str(nome).strip().lower().replace("ç", "c").replace("ã", "a").replace("á", "a")


def extrair_medicao_consolidada(arquivo):
    wb = openpyxl.load_workbook(arquivo, data_only=True)

    aba_alvo = None
    for nome in wb.sheetnames:
        normalizado = nome_aba_normalizado(nome)
        if "medicao consolidada" in normalizado:
            aba_alvo = nome
            break

    if not aba_alvo:
        raise ValueError("A aba 'MEDIÇÃO CONSOLIDADA' não foi encontrada na planilha.")

    ws = wb[aba_alvo]

    cabecalho = {
        "processo": ws["C13"].value,
        "rodovia": ws["C14"].value,
        "trecho": ws["C15"].value,
        "sub_trecho": ws["C16"].value,
        "segmento": ws["C17"].value,
        "extensao": ws["C18"].value,
        "aba_lida": aba_alvo,
    }

    itens = []

    # A planilha real termina os itens antes da linha 311 (que é total geral)
    for linha in range(23, 311):
        codigo_a = texto(ws.cell(linha, 1).value)   # coluna A
        codigo_b = texto(ws.cell(linha, 2).value)   # coluna B
        descricao = texto(ws.cell(linha, 3).value)  # coluna C
        unidade = texto(ws.cell(linha, 4).value)    # coluna D

        contrato_qtd = ws.cell(linha, 5).value      # E
        contrato_fin = ws.cell(linha, 6).value      # F

        qtd_acum_ant = ws.cell(linha, 18).value     # R
        qtd_liq_atual = ws.cell(linha, 19).value    # S
        qtd_acum_atual = ws.cell(linha, 20).value   # T

        preco_unit = ws.cell(linha, 21).value       # U

        fin_acum_ant = ws.cell(linha, 27).value     # AA
        fin_liq_atual = ws.cell(linha, 28).value    # AB
        perc_reajuste = ws.cell(linha, 29).value    # AC
        valor_reajuste = ws.cell(linha, 30).value   # AD
        fin_acum_atual = ws.cell(linha, 31).value   # AE

        saldo_qtd = ws.cell(linha, 32).value        # AF
        saldo_fin = ws.cell(linha, 33).value        # AG
        exec_percent = ws.cell(linha, 34).value     # AH

        # Corrige linhas quebradas como a 273, onde A veio "'" e o identificador útil está em B
        codigo = codigo_a
        if codigo in (None, "", "'"):
            codigo = codigo_b

        # Ignora linhas de grupo/seção
        # Ex.: "Transportes", "REPARO PROFUNDO", títulos de bloco etc.
        if not descricao:
            continue

        tem_estrutura_de_item = any([
            unidade,
            contrato_qtd is not None,
            preco_unit is not None,
            fin_acum_atual is not None,
            saldo_fin is not None,
        ])

        if not tem_estrutura_de_item:
            continue

        grupo = None
        if codigo:
            codigo_str = str(codigo).strip()
            if "." in codigo_str:
                grupo = codigo_str.split(".")[0]
            else:
                grupo = codigo_str

        itens.append({
            "linha": linha,
            "codigo": codigo,
            "codigo_auxiliar": codigo_b,
            "grupo": grupo,
            "item": descricao,
            "unidade": unidade,
            "contrato_quantidade": numero(contrato_qtd),
            "contrato_financeiro": numero(contrato_fin),
            "quantidade_acumulada_anterior": numero(qtd_acum_ant),
            "quantidade_liquida_atual": numero(qtd_liq_atual),
            "quantidade_acumulada_atual": numero(qtd_acum_atual),
            "preco_unitario": numero(preco_unit),
            "financeiro_acumulado_anterior": numero(fin_acum_ant),
            "financeiro_liquido_atual": numero(fin_liq_atual),
            "percentual_reajuste": numero(perc_reajuste),
            "valor_reajuste": numero(valor_reajuste),
            "financeiro_acumulado_atual": numero(fin_acum_atual),
            "saldo_quantidade": numero(saldo_qtd),
            "saldo_financeiro": numero(saldo_fin),
            "percentual_execucao": numero(exec_percent),
        })

    # Totais oficiais da própria planilha
    resumo = {
        "total_itens": len(itens),
        "total_preco_unitario": numero(ws["U311"].value),
        "total_financeiro_acumulado_anterior": numero(ws["AA311"].value),
        "total_financeiro_liquido_atual": numero(ws["AB311"].value),
        "total_financeiro_acumulado_atual": numero(ws["AE311"].value),
        "total_saldo_financeiro": numero(ws["AG311"].value),
        "pi": numero(ws["J316"].value) if isinstance(ws["J316"].value, (int, float)) else numero(ws["J317"].value),
        "reajuste_total": numero(ws["J317"].value) if isinstance(ws["J317"].value, (int, float)) else numero(ws["J318"].value),
        "pi_mais_reajuste": numero(ws["J318"].value),
    }

    # Se as células J316:J318 vierem como texto/valor deslocado, corrige lendo coluna J/K
    if resumo["pi"] == 0:
        resumo["pi"] = numero(ws["K316"].value)
    if resumo["reajuste_total"] == 0:
        resumo["reajuste_total"] = numero(ws["K317"].value)
    if resumo["pi_mais_reajuste"] == 0:
        resumo["pi_mais_reajuste"] = numero(ws["K318"].value)

    # Dashboard por grupo
    grupos = {}
    for item in itens:
        chave = item["grupo"] or "SEM_GRUPO"
        if chave not in grupos:
            grupos[chave] = {
                "grupo": chave,
                "quantidade_itens": 0,
                "contrato_financeiro": 0.0,
                "financeiro_liquido_atual": 0.0,
                "financeiro_acumulado_atual": 0.0,
                "saldo_financeiro": 0.0,
            }

        grupos[chave]["quantidade_itens"] += 1
        grupos[chave]["contrato_financeiro"] += item["contrato_financeiro"]
        grupos[chave]["financeiro_liquido_atual"] += item["financeiro_liquido_atual"]
        grupos[chave]["financeiro_acumulado_atual"] += item["financeiro_acumulado_atual"]
        grupos[chave]["saldo_financeiro"] += item["saldo_financeiro"]

    grupos_dashboard = sorted(
        grupos.values(),
        key=lambda x: x["financeiro_liquido_atual"],
        reverse=True
    )

    top_itens_atual = sorted(
        itens,
        key=lambda x: x["financeiro_liquido_atual"],
        reverse=True
    )[:10]

    return cabecalho, itens, resumo, grupos_dashboard, top_itens_atual